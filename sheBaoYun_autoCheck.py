import time
import openpyxl
import pyperclip
from PIL import ImageGrab
import pyautogui

import pygetwindow as gw


def get_window_pos_size_by_title(title):
    try:
        window = gw.getWindowsWithTitle(title)[0]
        window.restore()
        time.sleep(0.5)
        x = window.left
        y = window.top
        width = window.width
        height = window.height
        print("找到了窗口", title)

        return x, y, width, height
    except:
        print("找不到 窗口", title)


x, y, width, height = get_window_pos_size_by_title("社保云缴费")
print("窗口位置：({}, {})".format(x, y))
print("窗口大小：{}x{}".format(width, height))


# 定位坐标 并且将鼠标移动到该特征点的中心
def locate_image_and_move_to(img):
    locate = pyautogui.locateCenterOnScreen(img, region=(x, y, x + width, y + height))
    pyautogui.moveTo(locate.x, locate.y)
    time.sleep(0.3)


# def find(img):
#     exist = False
#     i = 1
#     while i <= 10:
#         exist = is_exist(img)
#         if exist:
#             return True
#     return False


def sereen_xiao_cheng_xu(id, name, result):
    im = ImageGrab.grab(bbox=(x, y, x + width, y + height))
    path = './result/{a}_{b}_{c}.png'.format(a=id, b=name, c=result)
    im.save(path)


def canbao_or_not(id, name):
    while True:
        if is_exist("./res/wei_jiao_fei.png"):
            sereen_xiao_cheng_xu(id, name, "未交费")
            # 点击返回

            locate_image_and_move_to('./res/fan_hui.png')
            pyautogui.click()
            pyautogui.moveTo(x + width, y + height)
            time.sleep(1)
            locate_image_and_move_to('./res/fan_hui.png')
            pyautogui.click()
            time.sleep(0.5)
            return "未交费"
        if is_exist("./res/chongfu_shenbao.png"):
            # 截图保存
            sereen_xiao_cheng_xu(id, name, "重复申报")
            # 点击确定
            locate_image_and_move_to("./res/que_ding.png")
            pyautogui.click()
            time.sleep(0.5)
            locate_image_and_move_to("./res/fan_hui.png")
            pyautogui.click()
            time.sleep(0.5)
            # 点击返回
            return "重复申报"
        if is_exist("./res/wu_deng_ji.png"):
            sereen_xiao_cheng_xu(id, name, "无登记参保信息")
            # 点击确定
            locate_image_and_move_to("./res/que_ding.png")
            pyautogui.click()
            time.sleep(0.5)
            return "无登记参保信息"


# 定位 特征 坐标 有则返回坐标 没有返回空
def locate(img):
    try:
        locate = pyautogui.locateCenterOnScreen(img, region=(x, y, x + width, y + height), grayscale=True)
        return locate
    except:
        return None


# 判断 特征点是否存在 存在返回true 不存在 返回false
def is_exist(img):
    try:
        locate = pyautogui.locateCenterOnScreen(img, region=(x, y, x + width, y + height))
        if locate is not None:
            return True
        else:
            return False
    except:
        return False


def search_user(id, name):
    # 判断 是否存在广告
    if is_exist('./res/close.png'):
        locate_image_and_move_to('./res/close.png')
        pyautogui.click()
        time.sleep(0.4)

    # 判断是否可以返回首页
    if is_exist('./res/shouye.png'):
        locate_image_and_move_to('./res/shouye.png')
        pyautogui.click()
        time.sleep(0.4)

    # 判断是否 存在缴费按钮
    if is_exist('./res/jiaofei_icon.png'):
        locate_image_and_move_to('./res/jiaofei_icon.png')
        pyautogui.click()

        # 定位到输入身份证
        for a in range(1, 10):
            if is_exist('./res/id_cart_input.png'):
                locate_image_and_move_to('./res/id_cart_input.png')
                pyautogui.click()
                time.sleep(0.2)
                pyperclip.copy(id)
                pyautogui.hotkey('ctrl', 'v')
                break
            else:
                time.sleep(0.5)

        pyautogui.scroll(-300)
        time.sleep(0.3)
        # 定位到姓名输入框
        locate_image_and_move_to('./res/name_input.png')
        pyautogui.click()
        time.sleep(0.2)
        pyperclip.copy(name)
        pyautogui.hotkey('ctrl', 'v')

        if is_exist("./res/next_step.png"):
            locate_image_and_move_to('./res/next_step.png')
            pyautogui.click()
            time.sleep(0.5)

        # 定位到下一步

        # 等待 2秒钟 判断结果
        time.sleep(3.5)

        if is_exist("./res/next_step2.png"):
            locate_image_and_move_to('./res/next_step2.png')
            pyautogui.click()

        # 判断用户 是否是已交 还是未交 还是重复参保
        is_canbao = canbao_or_not(id, name)
        return is_canbao
    else:
        print("没有找到缴费按钮")
        return 'ERROR'


# 配置
# excel 表格名称
excelName = "data.xlsx"
# 要跑第几页的数据
sheetIndex = "0"
# 开始行数
startLine = "2"

# 结束行数
endLine = "2"

# 身份证号在第几列 从0开始
idCol = 0

# 姓名在第几列 从0开始
nameCol = 1

# 读取excel 表格
workbook = openpyxl.load_workbook('./data/' + excelName)
sheets = workbook.worksheets
sheet1 = sheets[sheetIndex]
row_count = sheet1.max_row
print("最大行数", row_count)
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.append(['身份证号', '姓名', '是否缴费'])
new_workbook.save('./data/result.xlsx')

for i in range(2, row_count + 1):
    row_list_data = []
    for row in sheet1[i]:
        row_list_data.append(row.value)

    print('第', i - 1, '条数据', row_list_data)
    # 执行查询操作
    result = search_user(row_list_data[idCol], row_list_data[nameCol])
    # 判断返回结果 已缴费 还是未交费
    print("用户", row_list_data[idCol], '姓名', row_list_data[nameCol], "结果:", result)
    # 写入excel
    new_sheet.append([row_list_data[idCol], row_list_data[nameCol], result])
